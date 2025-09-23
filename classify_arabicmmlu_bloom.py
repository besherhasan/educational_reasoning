#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
End-to-end script: classify every Arabic MMLU question into Bloom's Taxonomy
levels with soft percentages using the strongest GPT model you have access to
(e.g., GPT-5). Saves progress after each question and auto-resumes.

Outputs:
- JSONL:  arabicmmlu_bloom_labels.jsonl  (saved after every row)
- XLSX:   arabicmmlu_bloom_labels.xlsx   (appended after every row)
- Checkpoint: bloom_checkpoint.json      (updated after every row)
- Optional snapshots (valid JSON array): arabicmmlu_bloom_labels.json (every N rows and on exit)

Dataset:
We load the aggregated "All/dev.csv" and "All/test.csv" splits directly.

Usage:
  export OPENAI_API_KEY=sk-...
  python classify_arabicmmlu_bloom.py \
      --model gpt-5 \
      --outdir ./out \
      --json-snapshot-every 200

Dependencies:
  pip install datasets openai pandas openpyxl
"""

import os
import sys
import json
import time
import math
import argparse
from typing import Dict, List, Any, Tuple
from pathlib import Path

import pandas as pd
from datasets import load_dataset
from openpyxl import Workbook, load_workbook

# OpenAI SDK v1.x style
#   pip install --upgrade openai
from openai import OpenAI
from openai import APIConnectionError, RateLimitError, APIError, InternalServerError


DEV_URL = "https://huggingface.co/datasets/MBZUAI/ArabicMMLU/resolve/main/All/dev.csv"
TEST_URL = "https://huggingface.co/datasets/MBZUAI/ArabicMMLU/resolve/main/All/test.csv"

BLOOM_KEYS = ["Remember", "Understand", "Apply", "Analyze", "Evaluate"]


def load_arabicmmlu_all() -> Dict[str, Any]:
    """
    Load ArabicMMLU 'All' split CSVs directly as two HF splits: dev and test.
    Columns include:
      ID, Source, Country, Group, Subject, Level, Question, Context,
      Answer Key, Option 1..5, is_few_shot
    """
    ds = load_dataset(
        "csv",
        data_files={"dev": DEV_URL, "test": TEST_URL},
        encoding="utf-8"
    )
    return ds


def make_client(api_key: str) -> OpenAI:
    return OpenAI(api_key=api_key)


def build_messages(question_text: str, options: List[str]) -> List[Dict[str, str]]:
    sys_msg = (
        "You are an expert educational assessor. "
        "Classify the cognitive skills required to answer a given question using Bloom's Taxonomy. "
        "Return ONLY a JSON object with exactly these keys and integer percentages that sum to 100: "
        "Remember, Understand, Apply, Analyze, Evaluate. "
        "Do not include any extra text."
    )

    parts = [f"Question:\n{question_text.strip()}"]
    if options:
        joined = "\n".join([f"- {o}" for o in options if o and str(o).strip()])
        if joined.strip():
            parts.append(f"Options:\n{joined}")

    user_msg = (
        "\n\n".join(parts)
        + "\n\n"
        + "Output JSON only with the five keys. Values must be integers in [0,100] and sum to 100."
    )

    return [
        {"role": "system", "content": sys_msg},
        {"role": "user", "content": user_msg},
    ]



from openai import OpenAI
from openai import (
    APIConnectionError, RateLimitError, APIError, InternalServerError, BadRequestError
)

def call_model_with_backoff(
    client: OpenAI,
    model: str,
    messages: list[dict],
    max_retries: int = 6,
    initial_delay: float = 2.0,
) -> str:
    """
    Calls chat.completions with smart fallbacks:
    - Try JSON mode + temperature=0.0 for deterministic output.
    - If the model rejects temperature, retry without it.
    - If it also rejects response_format, retry without JSON mode.
    Exponential backoff on transient errors.
    """
    attempt = 0
    delay = initial_delay
    allow_temperature = True           # first try with temperature=0.0
    allow_response_format = True       # first try with JSON mode

    while True:
        try:
            kwargs = dict(
                model=model,
                messages=messages,
            )
            if allow_temperature:
                kwargs["temperature"] = 0.0
            if allow_response_format:
                kwargs["response_format"] = {"type": "json_object"}

            resp = client.chat.completions.create(**kwargs)
            return resp.choices[0].message.content

        except BadRequestError as e:
            msg = str(e).lower()
            # Example: "Unsupported value: 'temperature' does not support 0.0..."
            if "temperature" in msg and "unsupported" in msg and allow_temperature:
                allow_temperature = False
                continue
            # Some models may not support response_format=json_object on chat.completions
            if "response_format" in msg and "unsupported" in msg and allow_response_format:
                allow_response_format = False
                continue
            # If it is a different 400, re-raise after saving progress upstream
            raise

        except (RateLimitError, APIConnectionError, APIError, InternalServerError):
            attempt += 1
            if attempt > max_retries:
                raise
            time.sleep(delay)
            delay = min(delay * 2, 60.0)


def extract_json_dict(s: str) -> Dict[str, Any]:
    """
    Parse model JSON. If there's stray text, try to locate the first {...} block.
    """
    s = (s or "").strip()
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        # fallback: find first JSON object substring
        start = s.find("{")
        end = s.rfind("}")
        if start != -1 and end != -1 and end > start:
            candidate = s[start : end + 1]
            return json.loads(candidate)
        raise


def normalize_distribution(d: Dict[str, Any]) -> Dict[str, int]:
    """
    Ensure all Bloom keys exist as non-negative integers and sum to 100.
    Strategy:
      - Coerce values to float, clip negatives to 0
      - If all zeros, fall back to Remember=100
      - Normalize to sum 100 by proportional scaling and integer rounding with remainder fix
    """
    vals = []
    for k in BLOOM_KEYS:
        v = d.get(k, 0)
        try:
            v = float(v)
        except Exception:
            v = 0.0
        if math.isnan(v) or v < 0:
            v = 0.0
        vals.append(v)

    total = sum(vals)
    if total <= 0:
        out = {k: 0 for k in BLOOM_KEYS}
        out["Remember"] = 100
        return out

    # Scale to 0..100 then round
    scaled = [v * 100.0 / total for v in vals]
    ints = [int(round(x)) for x in scaled]
    # Adjust rounding drift to sum exactly 100
    diff = 100 - sum(ints)
    if diff != 0:
        # distribute diff by tweaking largest remainders if positive, or largest ints if negative
        remainders = [s - i for s, i in zip(scaled, ints)]
        order = sorted(range(len(ints)), key=lambda i: remainders[i], reverse=(diff > 0))
        for i in order:
            if diff == 0:
                break
            # Prevent negatives
            if diff < 0 and ints[i] == 0:
                continue
            ints[i] += 1 if diff > 0 else -1
            diff += -1 if diff > 0 else 1

    out = {k: max(0, int(v)) for k, v in zip(BLOOM_KEYS, ints)}
    # Final safety: enforce exact 100 by correcting the largest bucket
    final_sum = sum(out.values())
    if final_sum != 100:
        # fix on the key with the largest value
        largest_key = max(out, key=out.get)
        out[largest_key] += (100 - final_sum)
    return out


def append_jsonl(path: Path, obj: Dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False))
        f.write("\n")


def snapshot_json_array(jsonl_path: Path, json_array_path: Path) -> None:
    """
    Create a valid JSON array snapshot from JSONL. Useful for tools that expect .json arrays.
    """
    out = []
    if jsonl_path.exists():
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    out.append(json.loads(line))
                except Exception:
                    continue
    with json_array_path.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)


def init_excel(xlsx_path: Path, header: List[str]) -> None:
    if xlsx_path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "labels"
    ws.append(header)
    wb.save(str(xlsx_path))


def append_excel_row(xlsx_path: Path, header: List[str], row: Dict[str, Any]) -> None:
    if not xlsx_path.exists():
        init_excel(xlsx_path, header)
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    ws.append([row.get(h, "") for h in header])
    wb.save(str(xlsx_path))


def load_checkpoint(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"processed_keys": []}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"processed_keys": []}


def save_checkpoint(path: Path, ckpt: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=2)


def already_processed_set(jsonl_path: Path) -> set:
    """
    Build a set of unique keys we have already emitted to JSONL.
    Unique key design: f"{split}:{ID}"
    """
    done = set()
    if not jsonl_path.exists():
        return done
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                k = rec.get("_ukey")
                if k:
                    done.add(k)
            except Exception:
                continue
    return done


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", type=str, default=os.getenv("OPENAI_MODEL", "gpt-5"),
                        help="OpenAI model name to use, e.g., gpt-5. You can also set OPENAI_MODEL.")
    parser.add_argument("--outdir", type=str, default="./out", help="Output directory.")
    parser.add_argument("--json-snapshot-every", type=int, default=200,
                        help="Create a valid JSON array snapshot every N rows.")
    parser.add_argument("--limit", type=int, default=0,
                        help="For quick tests, limit total rows processed. 0 means all.")
    args = parser.parse_args()

    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        print("ERROR: Please set OPENAI_API_KEY.", file=sys.stderr)
        sys.exit(1)

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    jsonl_path = outdir / "arabicmmlu_bloom_labels.jsonl"
    json_snapshot_path = outdir / "arabicmmlu_bloom_labels.json"
    xlsx_path = outdir / "arabicmmlu_bloom_labels.xlsx"
    ckpt_path = outdir / "bloom_checkpoint.json"

    # Prepare Excel header
    excel_header = [
        "split", "ID", "Group", "Subject", "Level", "Country",
        "Question",
        "Option 1", "Option 2", "Option 3", "Option 4", "Option 5",
        "Remember", "Understand", "Apply", "Analyze", "Evaluate"
    ]
    init_excel(xlsx_path, excel_header)

    # Load dataset
    print("Loading ArabicMMLU 'All' splits...")
    ds = load_arabicmmlu_all()
    total_rows = len(ds["dev"]) + len(ds["test"])
    print(f"Loaded: dev={len(ds['dev'])}, test={len(ds['test'])}, total={total_rows}")

    # Initialize OpenAI client
    client = make_client(api_key=api_key)

    # Resume: use JSONL as source of truth for completed records
    processed_keys = already_processed_set(jsonl_path)
    print(f"Found {len(processed_keys)} already-processed rows in JSONL. Resuming...")

    # Iterate both splits
    splits = [("dev", ds["dev"]), ("test", ds["test"])]
    processed_counter = 0
    appended_counter = 0

    try:
        for split_name, split_ds in splits:
            for row in split_ds:
                # Unique key based on split and ID (ID is string in CSV)
                qid = str(row.get("ID", "")).strip()
                ukey = f"{split_name}:{qid}"

                if ukey in processed_keys:
                    continue

                # Extract fields
                question = str(row.get("Question", "") or "").strip()
                options = []
                for k in ["Option 1", "Option 2", "Option 3", "Option 4", "Option 5"]:
                    v = row.get(k, "")
                    if v is not None and str(v).strip() and str(v).lower() != "null":
                        options.append(str(v))

                if not question:
                    # Skip empty question rows
                    processed_keys.add(ukey)
                    continue

                messages = build_messages(question_text=question, options=options)

                # Call the model with backoff
                raw = call_model_with_backoff(client, args.model, messages)

                # Parse and normalize
                parsed = extract_json_dict(raw)
                dist = normalize_distribution(parsed)

                # Build record
                out_rec = {
                    "_ukey": ukey,
                    "split": split_name,
                    "ID": qid,
                    "Group": row.get("Group", ""),
                    "Subject": row.get("Subject", ""),
                    "Level": row.get("Level", ""),
                    "Country": row.get("Country", ""),
                    "Question": question,
                    "Option 1": row.get("Option 1", ""),
                    "Option 2": row.get("Option 2", ""),
                    "Option 3": row.get("Option 3", ""),
                    "Option 4": row.get("Option 4", ""),
                    "Option 5": row.get("Option 5", ""),
                    "Remember": dist["Remember"],
                    "Understand": dist["Understand"],
                    "Apply": dist["Apply"],
                    "Analyze": dist["Analyze"],
                    "Evaluate": dist["Evaluate"],
                }

                # Save immediately to JSONL and XLSX
                append_jsonl(jsonl_path, out_rec)
                append_excel_row(xlsx_path, excel_header, out_rec)

                processed_keys.add(ukey)
                processed_counter += 1
                appended_counter += 1

                # Optional JSON array snapshot every N rows
                if args.json_snapshot_every and (appended_counter % args.json_snapshot_every == 0):
                    snapshot_json_array(jsonl_path, json_snapshot_path)

                # Respect optional limit for quick test runs
                if args.limit and processed_counter >= args.limit:
                    break

            if args.limit and processed_counter >= args.limit:
                break

    finally:
        # Final snapshot to a valid JSON array
        snapshot_json_array(jsonl_path, json_snapshot_path)

        # Save explicit checkpoint
        ckpt = {"processed_count": len(processed_keys), "processed_keys": sorted(list(processed_keys))}
        save_checkpoint(ckpt_path, ckpt)

    print(f"Done. Total newly processed rows: {appended_counter}")
    print(f"JSONL:  {jsonl_path}")
    print(f"XLSX:   {xlsx_path}")
    print(f"JSON:   {json_snapshot_path}")
    print(f"CKPT:   {ckpt_path}")


if __name__ == "__main__":
    main()
