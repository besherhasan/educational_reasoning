#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IndoMMLU → 2-way taxonomy (Memorization vs Reasoning, soft %) using OpenAI Batch API.

Key points in this version
- Two output keys only: `Memorization`, `Reasoning` (sum to 100).
- Unified `Options` list: stored as a JSON array in JSONL; stringified for Excel cells.
- Robust options parsing (list literals / newline / CSV) + label cleanup (A./B) ...).
- Prompts include the question + options.

Usage
  export OPENAI_API_KEY=sk-...
  python classify_indo_mmlu_bloom_batch.py \
      --model gpt-5 \
      --outdir ./out \
      --json-snapshot-every 200 \
      --batch-size 5000 \
      --poll-interval 30 \
      --max-wait-seconds 86400
  # quick sanity check:
  python classify_indo_mmlu_bloom_batch.py --debug --model gpt-5 --outdir ./out

Dependencies
  pip install datasets openai pandas openpyxl
"""

import os
import re
import ast
import sys
import json
import time
import math
import argparse
from typing import Dict, List, Any, Tuple
from pathlib import Path

import pandas as pd
from datasets import load_dataset
from openpyxl import Workbook, load_workbook

# OpenAI SDK v1.x
from openai import OpenAI
from openai import (
    APIConnectionError, RateLimitError, APIError, InternalServerError, BadRequestError
)

# ==== ONLY TWO KEYS NOW ====
BLOOM_KEYS = ["Memorization", "Reasoning"]

# -------------------- Dataset helpers --------------------

def load_indommlu_all() -> Dict[str, Any]:
    """Load IndoMMLU from Hugging Face."""
    return load_dataset("indolem/IndoMMLU")

def _parse_options_from_string(s: str) -> List[str]:
    """Parse options from various string formats (list literal / lines / CSV)."""
    s = (s or "").strip()

    # Try Python- or JSON-like list literal: "['A...', 'B...']" or '["A...", "B..."]'
    if s.startswith("[") and s.endswith("]"):
        # Try ast.literal_eval first (handles single quotes), then JSON
        for parser in (ast.literal_eval, json.loads):
            try:
                lst = parser(s)
                if isinstance(lst, list):
                    return [str(x).strip() for x in lst]
            except Exception:
                pass

    # Newline-separated
    if "\n" in s:
        return [p.strip() for p in s.split("\n") if p.strip()]

    # Comma-separated with labels (A./B)/...)
    if "," in s and any(lbl in s for lbl in ["A.", "B.", "C.", "D.", "E.", "A)", "B)", "C)", "D)", "E)"]):
        parts = re.split(r",\s*(?=[A-Ea-e][\.\)])", s)
        parts = [p.strip() for p in parts if p.strip()]
        return parts

    # Plain CSV as a last resort
    if "," in s:
        return [p.strip() for p in s.split(",") if p.strip()]

    # Nothing to split; treat the whole string as one option
    return [s] if s else []

_LABEL_RE = re.compile(r"^\s*([A-Ea-e])[\.\)]\s*")  # strips "A. " or "B) "

def _clean_option(o: str) -> str:
    o = str(o or "").strip()
    # Remove leading A./B)/...
    o = _LABEL_RE.sub("", o)
    return o

def extract_options_from_row(row: Dict[str, Any]) -> List[str]:
    """
    Pull options from:
      - dataset row 'options' (list OR string),
      - dataset row 'jawaban' (some variants),
      - or minimal row 'Options' (list OR string).
    Cleans labels and caps to 5 options.
    """
    src = None
    if "Options" in row:
        src = row["Options"]
    elif "options" in row:
        src = row["options"]
    elif "jawaban" in row:
        src = row["jawaban"]

    if isinstance(src, list):
        opts = [str(x) for x in src]
    elif isinstance(src, str):
        opts = _parse_options_from_string(src)
    else:
        # Last resort: legacy Option 1..5 style, if present
        tmp = []
        for i in range(1, 6):
            v = row.get(f"Option {i}")
            if v is not None and str(v).strip():
                tmp.append(str(v))
        opts = tmp

    cleaned = [_clean_option(o) for o in opts if str(o).strip()]
    return cleaned[:5]

# -------------------- OpenAI helpers --------------------

def make_client(api_key: str) -> OpenAI:
    return OpenAI(api_key=api_key)

def build_messages(question_text: str, options: List[str]) -> List[Dict[str, Any]]:
    sys_msg = (
        "You are an expert educational assessor. "
        "Classify the cognitive demand of answering a given multiple-choice question "
        "into exactly two categories with integer percentages that sum to 100: "
        "Memorization (recall/recognition of facts/definitions) and "
        "Reasoning (understanding, application, analysis, or evaluation beyond recall). "
        "Return ONLY a JSON object with exactly these keys: Memorization, Reasoning. "
        "Do not include any extra text."
    )
    parts = [f"Question:\n{question_text.strip()}"]
    if options:
        joined = "\n".join([f"- {o}" for o in options if o and str(o).strip()])
        if joined.strip():
            parts.append(f"Options:\n{joined}")
    user_msg = (
        "\n\n".join(parts)
        + "\n\n"
        + "Output JSON only with the two keys. Values must be integers in [0,100] and sum to 100."
    )
    return [
        {"role": "system", "content": sys_msg},
        {"role": "user", "content": user_msg},
    ]

def call_model_with_backoff(
    client: OpenAI,
    model: str,
    messages: List[Dict[str, Any]],
    max_retries: int = 6,
    initial_delay: float = 2.0,
) -> str:
    attempt = 0
    delay = initial_delay
    allow_temperature = True
    allow_response_format = True
    while True:
        try:
            kwargs = dict(model=model, messages=messages)
            if allow_temperature:
                kwargs["temperature"] = 0.0
            if allow_response_format:
                kwargs["response_format"] = {"type": "json_object"}
            resp = client.chat.completions.create(**kwargs)
            return resp.choices[0].message.content
        except BadRequestError as e:
            msg = str(e).lower()
            if "temperature" in msg and "unsupported" in msg and allow_temperature:
                allow_temperature = False
                continue
            if "response_format" in msg and "unsupported" in msg and allow_response_format:
                allow_response_format = False
                continue
            raise
        except (RateLimitError, APIConnectionError, APIError, InternalServerError):
            attempt += 1
            if attempt > max_retries:
                raise
            time.sleep(delay)
            delay = min(delay * 2, 60.0)

# -------------------- JSON/Excel helpers --------------------

def extract_json_dict(s: str) -> Dict[str, Any]:
    s = (s or "").strip()
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        start = s.find("{")
        end = s.rfind("}")
        if start != -1 and end != -1 and end > start:
            candidate = s[start : end + 1]
            return json.loads(candidate)
        raise

def normalize_distribution(d: Dict[str, Any]) -> Dict[str, int]:
    vals = []
    for k in BLOOM_KEYS:
        v = d.get(k, 0)
        try:
            v = float(v)
        except Exception:
            v = 0.0
        if math.isnan(v) or v < 0:
            v = 0.0
        vals.append(v)

    total = sum(vals)
    if total <= 0:
        out = {k: 0 for k in BLOOM_KEYS}
        out["Memorization"] = 100
        return out

    scaled = [v * 100.0 / total for v in vals]
    ints = [int(round(x)) for x in scaled]
    diff = 100 - sum(ints)
    if diff != 0:
        remainders = [s - i for s, i in zip(scaled, ints)]
        order = sorted(range(len(ints)), key=lambda i: remainders[i], reverse=(diff > 0))
        for i in order:
            if diff == 0:
                break
            if diff < 0 and ints[i] == 0:
                continue
            ints[i] += 1 if diff > 0 else -1
            diff += -1 if diff > 0 else 1

    out = {k: max(0, int(v)) for k, v in zip(BLOOM_KEYS, ints)}
    final_sum = sum(out.values())
    if final_sum != 100:
        largest_key = max(out, key=out.get)
        out[largest_key] += (100 - final_sum)
    return out

def excel_safe_value(value: Any) -> Any:
    """
    Convert Python objects into values acceptable by openpyxl cells.
    - Lists/Dicts -> JSON strings (UTF-8, not ASCII-escaped).
    - Numpy scalars -> Python scalars.
    - None -> "".
    Also truncates values to Excel's 32,767-character cell limit.
    """
    if value is None:
        s = ""
    elif isinstance(value, (list, dict)):
        s = json.dumps(value, ensure_ascii=False)
    else:
        # Convert numpy scalars if present
        try:
            module = type(value).__module__
            if module and module.startswith("numpy"):  # e.g., numpy.int64
                try:
                    value = value.item()
                except Exception:
                    value = str(value)
        except Exception:
            pass
        s = value

    # Enforce Excel cell size limit
    if isinstance(s, str) and len(s) > 32760:
        s = s[:32760] + "…"
    return s

def append_jsonl(path: Path, obj: Dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False))
        f.write("\n")

def snapshot_json_array(jsonl_path: Path, json_array_path: Path) -> None:
    out = []
    if jsonl_path.exists():
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    out.append(json.loads(line))
                except Exception:
                    continue
    with json_array_path.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

def init_excel(xlsx_path: Path, header: List[str]) -> None:
    if xlsx_path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "labels"
    ws.append(header)
    wb.save(str(xlsx_path))

def append_excel_row(xlsx_path: Path, header: List[str], row: Dict[str, Any]) -> None:
    if not xlsx_path.exists():
        init_excel(xlsx_path, header)
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    values = [excel_safe_value(row.get(h, "")) for h in header]
    ws.append(values)
    wb.save(str(xlsx_path))

def load_checkpoint(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"processed_keys": []}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"processed_keys": []}

def save_checkpoint(path: Path, ckpt: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=2)

def already_processed_set(jsonl_path: Path) -> set:
    done = set()
    if not jsonl_path.exists():
        return done
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                k = rec.get("_ukey")
                if k:
                    done.add(k)
            except Exception:
                continue
    return done

# -------------------- Batch helpers --------------------

def build_batch_tasks(
    pending_rows: List[Tuple[str, Dict[str, Any]]],
    model: str
) -> List[Dict[str, Any]]:
    tasks = []
    for ukey, row in pending_rows:
        question = str(row.get("Question") or row.get("question") or "").strip()
        if not question:
            continue
        options = extract_options_from_row(row)

        messages = build_messages(question, options)
        body = {"model": model, "messages": messages}

        task = {
            "custom_id": ukey,
            "method": "POST",
            "url": "/v1/chat/completions",
            "body": body
        }
        tasks.append(task)
    return tasks

def write_jsonl_tasks(tasks: List[Dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for obj in tasks:
            f.write(json.dumps(obj, ensure_ascii=False))
            f.write("\n")

def submit_and_wait_batch(client: OpenAI, file_id: str, endpoint: str, poll_interval: int, max_wait_seconds: int) -> Dict[str, Any]:
    job = client.batches.create(
        input_file_id=file_id,
        endpoint=endpoint,
        completion_window="24h"
    )
    start = time.time()
    while True:
        job = client.batches.retrieve(job.id)
        status = getattr(job, "status", None)
        if status in ("completed", "failed", "expired", "cancelling", "cancelled"):
            return job
        if time.time() - start > max_wait_seconds:
            return job
        time.sleep(poll_interval)

def download_batch_results(client: OpenAI, output_file_id: str, out_path: Path) -> None:
    content = client.files.content(output_file_id).content
    with out_path.open("wb") as f:
        f.write(content)

def iter_batch_result_lines(result_jsonl_path: Path):
    with result_jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except Exception:
                continue

# -------------------- Main --------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", type=str, default=os.getenv("OPENAI_MODEL", "gpt-5"))
    parser.add_argument("--outdir", type=str, default="./out")
    parser.add_argument("--json-snapshot-every", type=int, default=200)
    parser.add_argument("--limit", type=int, default=0, help="Limit rows processed. 0 means all.")
    parser.add_argument("--debug", action="store_true", help="Run on 5 samples synchronously without Batch.")
    parser.add_argument("--batch-size", type=int, default=5000, help="Max requests per batch file.")
    parser.add_argument("--poll-interval", type=int, default=30, help="Seconds between batch status polls.")
    parser.add_argument("--max-wait-seconds", type=int, default=86400, help="Max total wait in seconds.")
    parser.add_argument("--endpoint", type=str, default="/v1/chat/completions")
    args = parser.parse_args()

    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        print("ERROR: Please set OPENAI_API_KEY.", file=sys.stderr)
        sys.exit(1)

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    jsonl_path = outdir / "indommlu_bloom_labels.jsonl"
    json_snapshot_path = outdir / "indommlu_bloom_labels.json"
    xlsx_path = outdir / "indommlu_bloom_labels.xlsx"
    ckpt_path = outdir / "bloom_checkpoint.json"

    # Excel columns updated for two keys
    excel_header = [
        "split", "ID", "Subject", "Level",
        "Question",
        "Options",
        "Memorization", "Reasoning"
    ]
    init_excel(xlsx_path, excel_header)

    print("Loading IndoMMLU 'test' split...")
    ds = load_indommlu_all()
    total_rows = len(ds["test"])
    print(f"Loaded: test={total_rows}")

    client = make_client(api_key=api_key)

    processed_keys = already_processed_set(jsonl_path)
    print(f"Found {len(processed_keys)} already-processed rows in JSONL. Resuming...")

    # Build list of pending rows with unique keys in stable order
    splits = [("test", ds["test"])]
    pending: List[Tuple[str, Dict[str, Any]]] = []
    for split_name, split_ds in splits:
        for idx, row in enumerate(split_ds):
            ukey = f"{split_name}:{idx}"
            if ukey in processed_keys:
                continue
            # Skip empty questions early
            if not str(row.get("question", "") or "").strip():
                continue

            # Extract options and store under unified "Options"
            opts = extract_options_from_row(row)

            # Keep minimal fields for downstream
            minimal = {
                "split": split_name,
                "ID": str(idx),
                "Subject": row.get("subject", ""),
                "Level": row.get("level", ""),
                "Question": row.get("question", ""),
                "Options": opts,  # list for JSONL; stringified for Excel
            }
            pending.append((ukey, minimal))
            if args.limit and len(pending) >= args.limit:
                break
        if args.limit and len(pending) >= args.limit:
            break

    if args.debug:
        # Synchronous sanity check on up to 5 rows
        debug_rows = pending[:5]
        appended_counter = 0
        for ukey, row in debug_rows:
            options = extract_options_from_row(row)
            messages = build_messages(row["Question"], options)
            raw = call_model_with_backoff(client, args.model, messages)
            parsed = extract_json_dict(raw)
            dist = normalize_distribution(parsed)
            out_rec = {
                "_ukey": ukey,
                **row,  # includes "Options" list
                "Memorization": dist["Memorization"],
                "Reasoning": dist["Reasoning"],
            }
            append_jsonl(jsonl_path, out_rec)
            append_excel_row(xlsx_path, excel_header, out_rec)
            processed_keys.add(ukey)
            appended_counter += 1

        snapshot_json_array(jsonl_path, json_snapshot_path)
        ckpt = {
            "processed_count": len(processed_keys),
            "processed_keys": sorted(list(processed_keys))
        }
        save_checkpoint(ckpt_path, ckpt)
        print(f"Debug done. Newly processed rows: {appended_counter}")
        print(f"JSONL:  {jsonl_path}")
        print(f"XLSX:   {xlsx_path}")
        print(f"JSON:   {json_snapshot_path}")
        print(f"CKPT:   {ckpt_path}")
        return

    if not pending:
        print("Nothing to process. All rows are already labeled.")
        return

    # Chunk into batch files
    batch_size = max(1, int(args.batch_size))
    chunks = [pending[i:i+batch_size] for i in range(0, len(pending), batch_size)]
    all_jobs = []

    for bi, chunk_rows in enumerate(chunks, start=1):
        print(f"Preparing batch {bi}/{len(chunks)} with {len(chunk_rows)} requests...")
        tasks = build_batch_tasks(chunk_rows, args.model)
        if not tasks:
            continue
        batch_tasks_path = outdir / f"batch_tasks_{bi:04d}.jsonl"
        write_jsonl_tasks(tasks, batch_tasks_path)

        # Upload batch file
        print(f"Uploading {batch_tasks_path.name}...")
        batch_file = client.files.create(file=open(batch_tasks_path, "rb"), purpose="batch")
        print(f"Uploaded file_id={batch_file.id}, bytes={batch_file.bytes}, status={batch_file.status}")

        # Create and wait for batch job
        print("Submitting batch job...")
        job = submit_and_wait_batch(
            client=client,
            file_id=batch_file.id,
            endpoint=args.endpoint,
            poll_interval=args.poll_interval,
            max_wait_seconds=args.max_wait_seconds
        )
        print(f"Batch job status: {getattr(job, 'status', None)}  id={job.id}")

        # Save intermediate checkpoint
        ckpt = {"processed_count": len(processed_keys),
                "processed_keys": sorted(list(processed_keys)),
                "last_batch_id": job.id}
        save_checkpoint(ckpt_path, ckpt)

        # If job has output, download and ingest
        output_file_id = getattr(job, "output_file_id", None)
        if output_file_id:
            result_path = outdir / f"batch_results_{bi:04d}.jsonl"
            print(f"Downloading results to {result_path.name}...")
            download_batch_results(client, output_file_id, result_path)

            # Build quick index map for this chunk (ukey -> row)
            source_map = {k: v for k, v in chunk_rows}

            # Ingest results line by line and persist per row
            ingested = 0
            for res in iter_batch_result_lines(result_path):
                try:
                    ukey = res.get("custom_id")
                    if not ukey:
                        continue

                    body = (((res.get("response") or {}).get("body") or {}))
                    choices = (body.get("choices") or [])
                    content = ""
                    if choices and "message" in choices[0]:
                        content = choices[0]["message"].get("content", "")
                    elif "output" in body:
                        content = body["output"]
                    if not content:
                        continue

                    row = source_map.get(ukey)
                    if not row:
                        continue

                    parsed = extract_json_dict(content)
                    dist = normalize_distribution(parsed)

                    out_rec = {
                        "_ukey": ukey,
                        **row,  # includes "Options" list
                        "Memorization": dist["Memorization"],
                        "Reasoning": dist["Reasoning"],
                    }

                    append_jsonl(jsonl_path, out_rec)
                    append_excel_row(xlsx_path, excel_header, out_rec)
                    processed_keys.add(ukey)
                    ingested += 1

                    if args.json_snapshot_every and (ingested % args.json_snapshot_every == 0):
                        snapshot_json_array(jsonl_path, json_snapshot_path)
                        ckpt = {"processed_count": len(processed_keys),
                                "processed_keys": sorted(list(processed_keys))}
                        save_checkpoint(ckpt_path, ckpt)
                except Exception as e:
                    # Continue robustly and record a minimal failure note
                    fail_note = {"_ukey": res.get("custom_id", ""), "error": str(e)[:400]}
                    append_jsonl(jsonl_path, fail_note)

            # Finalize per batch
            snapshot_json_array(jsonl_path, json_snapshot_path)
            ckpt = {"processed_count": len(processed_keys),
                    "processed_keys": sorted(list(processed_keys))}
            save_checkpoint(ckpt_path, ckpt)
            print(f"Ingested {ingested} rows from batch {bi}.")

        all_jobs.append(job.id)

    print(f"Done. Batches submitted: {len(all_jobs)}")
    print(f"JSONL:  {jsonl_path}")
    print(f"XLSX:   {xlsx_path}")
    print(f"JSON:   {json_snapshot_path}")
    print(f"CKPT:   {ckpt_path}")

if __name__ == "__main__":
    main()
