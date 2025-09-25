#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
End-to-end script: classify every MMLU (English) question into Bloom's Taxonomy
(5-way: Remember, Understand, Apply, Analyze, Evaluate) with soft percentages
using OpenAI's Batch API.

Outputs
- JSONL:  mmlu_reasoning_labels.jsonl  (written per row on result ingest)
- XLSX:   mmlu_reasoning_labels.xlsx   (appended per row on result ingest)
- Checkpoint: mmlu_reasoning_checkpoint.json  (updated on every ingest or poll)
- Optional snapshots: mmlu_reasoning_labels.json (valid JSON array)

Usage
  export OPENAI_API_KEY=sk-...
  python classify_mmlu_reasoning_batch.py \
      --model gpt-5 \
      --outdir ./out \
      --json-snapshot-every 200 \
      --batch-size 5000 \
      --poll-interval 30 \
      --max-wait-seconds 86400

Dependencies
  pip install datasets openai pandas openpyxl
"""

import os
import sys
import json
import time
import math
import argparse
from typing import Dict, List, Any, Tuple
from pathlib import Path

import pandas as pd  # kept for parity, not strictly required
from datasets import load_dataset
from openpyxl import Workbook, load_workbook

# OpenAI SDK v1.x
from openai import OpenAI
from openai import (
    APIConnectionError, RateLimitError, APIError, InternalServerError, BadRequestError
)

# ---- Five-way Bloom's levels ----
BLOOM_KEYS = ["Remember", "Understand", "Apply", "Analyze", "Evaluate"]


def load_mmlu_all() -> Dict[str, Any]:
    """
    Load full English MMLU with all subjects using the 'all' config.
    Expected columns per row:
      - question: str
      - subject: str
      - choices: list[str] of length 4
      - answer: int (0..3)
    """
    ds = load_dataset("cais/mmlu", "all")
    return ds


def make_client(api_key: str) -> OpenAI:
    return OpenAI(api_key=api_key)


def build_messages(question_text: str, options: List[str]) -> List[Dict[str, Any]]:
    # << your requested prompt exactly, adapted for this script >>
    sys_msg = (
        "You are an expert educational assessor. "
        "Classify the cognitive skills required to answer a given question using Bloom's Taxonomy. "
        "Return ONLY a JSON object with exactly these keys and integer percentages that sum to 100: "
        "Remember, Understand, Apply, Analyze, Evaluate. "
        "Do not include any extra text."
    )
    parts = [f"Question:\n{question_text.strip()}"]
    if options:
        joined = "\n".join([f"- {o}" for o in options if o and str(o).strip()])
        if joined.strip():
            parts.append(f"Options:\n{joined}")
    user_msg = (
        "\n\n".join(parts)
        + "\n\n"
        + "Output JSON only with the five keys. Values must be integers in [0,100] and sum to 100."
    )
    return [
        {"role": "system", "content": sys_msg},
        {"role": "user", "content": user_msg},
    ]


def call_model_with_backoff(
    client: OpenAI,
    model: str,
    messages: List[Dict[str, Any]],
    max_retries: int = 6,
    initial_delay: float = 2.0,
) -> str:
    attempt = 0
    delay = initial_delay
    allow_temperature = True
    allow_response_format = True
    while True:
        try:
            kwargs = dict(model=model, messages=messages)
            if allow_temperature:
                kwargs["temperature"] = 0.0
            if allow_response_format:
                kwargs["response_format"] = {"type": "json_object"}
            resp = client.chat.completions.create(**kwargs)
            return resp.choices[0].message.content
        except BadRequestError as e:
            msg = str(e).lower()
            if "temperature" in msg and "unsupported" in msg and allow_temperature:
                allow_temperature = False
                continue
            if "response_format" in msg and "unsupported" in msg and allow_response_format:
                allow_response_format = False
                continue
            raise
        except (RateLimitError, APIConnectionError, APIError, InternalServerError):
            attempt += 1
            if attempt > max_retries:
                raise
            time.sleep(delay)
            delay = min(delay * 2, 60.0)


def extract_json_dict(s: str) -> Dict[str, Any]:
    s = (s or "").strip()
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        start = s.find("{")
        end = s.rfind("}")
        if start != -1 and end != -1 and end > start:
            candidate = s[start : end + 1]
            return json.loads(candidate)
        raise


def normalize_distribution(d: Dict[str, Any]) -> Dict[str, int]:
    vals = []
    for k in BLOOM_KEYS:
        v = d.get(k, 0)
        try:
            v = float(v)
        except Exception:
            v = 0.0
        if math.isnan(v) or v < 0:
            v = 0.0
        vals.append(v)

    total = sum(vals)
    if total <= 0:
        out = {k: 0 for k in BLOOM_KEYS}
        out["Remember"] = 100
        return out

    scaled = [v * 100.0 / total for v in vals]
    ints = [int(round(x)) for x in scaled]
    diff = 100 - sum(ints)
    if diff != 0:
        remainders = [s - i for s, i in zip(scaled, ints)]
        order = sorted(range(len(ints)), key=lambda i: remainders[i], reverse=(diff > 0))
        for i in order:
            if diff == 0:
                break
            if diff < 0 and ints[i] == 0:
                continue
            ints[i] += 1 if diff > 0 else -1
            diff += -1 if diff > 0 else 1

    out = {k: max(0, int(v)) for k, v in zip(BLOOM_KEYS, ints)}
    final_sum = sum(out.values())
    if final_sum != 100:
        largest_key = max(out, key=out.get)
        out[largest_key] += (100 - final_sum)
    return out


def append_jsonl(path: Path, obj: Dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False))
        f.write("\n")


def snapshot_json_array(jsonl_path: Path, json_array_path: Path) -> None:
    out = []
    if jsonl_path.exists():
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    out.append(json.loads(line))
                except Exception:
                    continue
    with json_array_path.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)


def init_excel(xlsx_path: Path, header: List[str]) -> None:
    if xlsx_path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "labels"
    ws.append(header)
    wb.save(str(xlsx_path))


def append_excel_row(xlsx_path: Path, header: List[str], row: Dict[str, Any]) -> None:
    if not xlsx_path.exists():
        init_excel(xlsx_path, header)
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    ws.append([row.get(h, "") for h in header])
    wb.save(str(xlsx_path))


def load_checkpoint(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"processed_keys": []}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"processed_keys": []}


def save_checkpoint(path: Path, ckpt: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=2)


def already_processed_set(jsonl_path: Path) -> set:
    done = set()
    if not jsonl_path.exists():
        return done
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                k = rec.get("_ukey")
                if k:
                    done.add(k)
            except Exception:
                continue
    return done


# ---------- Batch helpers ----------

def build_batch_tasks(
    pending_rows: List[Tuple[str, Dict[str, Any]]],
    model: str
) -> List[Dict[str, Any]]:
    tasks = []
    for ukey, row in pending_rows:
        question = str(row.get("Question", "") or "").strip()
        options = []
        for k in ["Option 1", "Option 2", "Option 3", "Option 4", "Option 5"]:
            v = row.get(k, "")
            if v is not None and str(v).strip() and str(v).lower() != "null":
                options.append(str(v))
        if not question:
            continue
        messages = build_messages(question, options)
        body = {"model": model, "messages": messages}

        task = {
            "custom_id": ukey,
            "method": "POST",
            "url": "/v1/chat/completions",
            "body": body
        }
        tasks.append(task)
    return tasks


def write_jsonl_tasks(tasks: List[Dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for obj in tasks:
            f.write(json.dumps(obj, ensure_ascii=False))
            f.write("\n")


def submit_and_wait_batch(client: OpenAI, file_id: str, endpoint: str, poll_interval: int, max_wait_seconds: int) -> Dict[str, Any]:
    job = client.batches.create(
        input_file_id=file_id,
        endpoint=endpoint,
        completion_window="24h"
    )
    start = time.time()
    while True:
        job = client.batches.retrieve(job.id)
        status = getattr(job, "status", None)
        if status in ("completed", "failed", "expired", "cancelling", "cancelled"):
            return job
        if time.time() - start > max_wait_seconds:
            return job
        time.sleep(poll_interval)


def download_batch_results(client: OpenAI, output_file_id: str, out_path: Path) -> None:
    content = client.files.content(output_file_id).content
    with out_path.open("wb") as f:
        f.write(content)


def iter_batch_result_lines(result_jsonl_path: Path):
    with result_jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except Exception:
                continue


# ---------- Main ----------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", type=str, default=os.getenv("OPENAI_MODEL", "gpt-5"))
    parser.add_argument("--outdir", type=str, default="./out")
    parser.add_argument("--json-snapshot-every", type=int, default=200)
    parser.add_argument("--limit", type=int, default=0, help="Limit rows processed. 0 means all.")
    parser.add_argument("--debug", action="store_true", help="Run on 5 samples synchronously without Batch.")
    parser.add_argument("--batch-size", type=int, default=5000, help="Max requests per batch file.")
    parser.add_argument("--poll-interval", type=int, default=30, help="Seconds between batch status polls.")
    parser.add_argument("--max-wait-seconds", type=int, default=86400, help="Max total wait in seconds.")
    parser.add_argument("--endpoint", type=str, default="/v1/chat/completions")
    args = parser.parse_args()

    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        print("ERROR: Please set OPENAI_API_KEY.", file=sys.stderr)
        sys.exit(1)

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    jsonl_path = outdir / "mmlu_reasoning_labels.jsonl"
    json_snapshot_path = outdir / "mmlu_reasoning_labels.json"
    xlsx_path = outdir / "mmlu_reasoning_labels.xlsx"
    ckpt_path = outdir / "mmlu_reasoning_checkpoint.json"

    # Update Excel header to 5 Bloom keys
    excel_header = [
        "split", "ID", "Subject", "Question",
        "Option 1", "Option 2", "Option 3", "Option 4",
        "Answer",
        "Remember", "Understand", "Apply", "Analyze", "Evaluate"
    ]
    init_excel(xlsx_path, excel_header)

    print("Loading MMLU 'all' splits...")
    ds = load_mmlu_all()
    split_names = list(ds.keys())
    total_rows = sum(len(ds[name]) for name in split_names)
    print(f"Loaded splits: {', '.join(split_names)}; total={total_rows}")

    client = make_client(api_key=api_key)

    processed_keys = already_processed_set(jsonl_path)
    print(f"Found {len(processed_keys)} already-processed rows in JSONL. Resuming...")

    # Build list of pending rows with stable unique keys
    pending: List[Tuple[str, Dict[str, Any]]] = []
    for split_name in split_names:
        split_ds = ds[split_name]
        for i, row in enumerate(split_ds):
            # MMLU schema
            q = row.get("question", "")
            choices = row.get("choices", None)
            subj = row.get("subject", "")
            ans = row.get("answer", "")

            if not str(q or "").strip():
                continue

            # Unique key per row
            ukey = f"{split_name}:{i}"
            if ukey in processed_keys:
                continue

            # Map choices to Option 1..4
            o1 = choices[0] if choices and len(choices) > 0 else ""
            o2 = choices[1] if choices and len(choices) > 1 else ""
            o3 = choices[2] if choices and len(choices) > 2 else ""
            o4 = choices[3] if choices and len(choices) > 3 else ""

            minimal = {
                "split": split_name,
                "ID": str(i),
                "Subject": subj,
                "Question": q,
                "Option 1": o1,
                "Option 2": o2,
                "Option 3": o3,
                "Option 4": o4,
                "Answer": ans,
            }
            pending.append((ukey, minimal))
            if args.limit and len(pending) >= args.limit:
                break
        if args.limit and len(pending) >= args.limit:
            break

    if args.debug:
        debug_rows = pending[:5]
        appended_counter = 0
        for ukey, row in debug_rows:
            messages = build_messages(
                row["Question"],
                [row.get(f"Option {i}") for i in range(1, 5)]
            )
            raw = call_model_with_backoff(client, args.model, messages)
            parsed = extract_json_dict(raw)
            dist = normalize_distribution(parsed)
            out_rec = {
                "_ukey": ukey,
                **row,
                "Remember": dist["Remember"],
                "Understand": dist["Understand"],
                "Apply": dist["Apply"],
                "Analyze": dist["Analyze"],
                "Evaluate": dist["Evaluate"],
            }
            append_jsonl(jsonl_path, out_rec)
            append_excel_row(xlsx_path, excel_header, out_rec)
            processed_keys.add(ukey)
            appended_counter += 1

        snapshot_json_array(jsonl_path, json_snapshot_path)
        ckpt = {"processed_count": len(processed_keys), "processed_keys": sorted(list(processed_keys))}
        save_checkpoint(ckpt_path, ckpt)
        print(f"Debug done. Newly processed rows: {appended_counter}")
        print(f"JSONL:  {jsonl_path}")
        print(f"XLSX:   {xlsx_path}")
        print(f"JSON:   {json_snapshot_path}")
        print(f"CKPT:   {ckpt_path}")
        return

    if not pending:
        print("Nothing to process. All rows are already labeled.")
        return

    # Chunk into batch files
    batch_size = max(1, int(args.batch_size))
    chunks = [pending[i:i+batch_size] for i in range(0, len(pending), batch_size)]
    all_jobs = []

    for bi, chunk_rows in enumerate(chunks, start=1):
        print(f"Preparing batch {bi}/{len(chunks)} with {len(chunk_rows)} requests...")
        tasks = build_batch_tasks(chunk_rows, args.model)
        if not tasks:
            continue
        batch_tasks_path = outdir / f"batch_tasks_{bi:04d}.jsonl"
        write_jsonl_tasks(tasks, batch_tasks_path)

        # Upload batch file
        print(f"Uploading {batch_tasks_path.name}...")
        batch_file = client.files.create(file=open(batch_tasks_path, "rb"), purpose="batch")
        print(f"Uploaded file_id={batch_file.id}, bytes={batch_file.bytes}, status={batch_file.status}")

        # Create and wait for batch job
        print("Submitting batch job...")
        job = submit_and_wait_batch(
            client=client,
            file_id=batch_file.id,
            endpoint=args.endpoint,
            poll_interval=args.poll_interval,
            max_wait_seconds=args.max_wait_seconds
        )
        print(f"Batch job status: {getattr(job, 'status', None)}  id={job.id}")

        # Save intermediate checkpoint
        ckpt = {"processed_count": len(processed_keys), "processed_keys": sorted(list(processed_keys)), "last_batch_id": job.id}
        save_checkpoint(ckpt_path, ckpt)

        # If job has output, download and ingest
        output_file_id = getattr(job, "output_file_id", None)
        if output_file_id:
            result_path = outdir / f"batch_results_{bi:04d}.jsonl"
            print(f"Downloading results to {result_path.name}...")
            download_batch_results(client, output_file_id, result_path)

            # Ingest results line by line and persist per row
            ingested = 0
            source_map = {k: v for k, v in chunk_rows}

            for res in iter_batch_result_lines(result_path):
                try:
                    ukey = res.get("custom_id")
                    if not ukey:
                        continue
                    body = (((res.get("response") or {}).get("body") or {}))
                    choices = (body.get("choices") or [])
                    content = ""
                    if choices and "message" in choices[0]:
                        content = choices[0]["message"].get("content", "")
                    elif "output" in body:
                        content = body["output"]
                    if not content:
                        continue

                    row = source_map.get(ukey)
                    if not row:
                        continue

                    parsed = extract_json_dict(content)
                    dist = normalize_distribution(parsed)

                    out_rec = {
                        "_ukey": ukey,
                        **row,
                        "Remember": dist["Remember"],
                        "Understand": dist["Understand"],
                        "Apply": dist["Apply"],
                        "Analyze": dist["Analyze"],
                        "Evaluate": dist["Evaluate"],
                    }

                    append_jsonl(jsonl_path, out_rec)
                    append_excel_row(xlsx_path, excel_header, out_rec)
                    processed_keys.add(ukey)
                    ingested += 1

                    if args.json_snapshot_every and (ingested % args.json_snapshot_every == 0):
                        snapshot_json_array(jsonl_path, json_snapshot_path)
                        ckpt = {"processed_count": len(processed_keys), "processed_keys": sorted(list(processed_keys))}
                        save_checkpoint(ckpt_path, ckpt)
                except Exception as e:
                    fail_note = {"_ukey": res.get("custom_id", ""), "error": str(e)[:400]}
                    append_jsonl(jsonl_path, fail_note)

            # Finalize per batch
            snapshot_json_array(jsonl_path, json_snapshot_path)
            ckpt = {"processed_count": len(processed_keys), "processed_keys": sorted(list(processed_keys))}
            save_checkpoint(ckpt_path, ckpt)
            print(f"Ingested { ingested } rows from batch {bi}.")

        all_jobs.append(job.id)

    print(f"Done. Batches submitted: {len(all_jobs)}")
    print(f"JSONL:  {jsonl_path}")
    print(f"XLSX:   {xlsx_path}")
    print(f"JSON:   {json_snapshot_path}")
    print(f"CKPT:   {ckpt_path}")


if __name__ == "__main__":
    main()
